"""団体戦Excelパーサー: openpyxl でシートを読み、立教の成績を抽出する。"""

import re
from pathlib import Path

import openpyxl

from common import is_rikkyo, logger
from parse_team import DIVISION_RE, PROMOTION_KEYWORDS, _normalize_division


def parse_team_xlsx(xlsx_path: Path, event_name: str) -> dict | None:
    """
    団体戦 Excel を解析し、立教の成績を返す。
    立教が出場していない場合は None を返す。
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        division = None

        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            cells = [c for c in cells if c]  # 空セル除去

            if not cells:
                continue

            joined = " ".join(cells)

            # ヘッダー行: division 更新
            if "勝点" in joined or "勝数" in joined:
                m = DIVISION_RE.search(joined)
                if m:
                    division = _normalize_division(m.group(1))
                continue

            # 立教行
            if not is_rikkyo(joined):
                continue

            # 立教大学セルの位置を特定
            rikkyo_idx = next((i for i, c in enumerate(cells) if is_rikkyo(c)), 0)
            relevant = cells[rikkyo_idx + 1:]  # 名前セルの右側

            # 数値のみ抽出 + 複合トークン(「8降」等)を分離
            clean_nums = []
            promotion = None
            rank_raw = None

            for token in relevant:
                t = token.strip()
                if re.match(r"^\d+$", t):
                    clean_nums.append(int(t))
                elif re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", t):
                    mp = re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", t)
                    rank_raw = int(mp.group(1))
                    promotion = "昇級" if "昇" in mp.group(2) else "降級"
                elif t in PROMOTION_KEYWORDS:
                    promotion = t

            try:
                if rank_raw is not None:
                    rank = rank_raw
                    points = clean_nums[-1] if clean_nums else None
                    wins_matches = clean_nums[-2] if len(clean_nums) >= 2 else None
                else:
                    rank = clean_nums[-1]
                    points = clean_nums[-2] if len(clean_nums) >= 2 else None
                    wins_matches = clean_nums[-3] if len(clean_nums) >= 3 else None
            except (IndexError, ValueError):
                logger.warning("Excel立教行の解析失敗: %s", cells[:8])
                continue

            logger.info("立教発見(XLSX): %s %s rank=%s wins=%s points=%s promo=%s",
                        event_name, division or "?", rank, wins_matches, points, promotion)
            return {"division": division, "result": {
                "rank": rank,
                "wins": wins_matches,
                "losses": None,
                "points": points,
                "promotion": promotion,
                "note": "",
            }}

    return None
