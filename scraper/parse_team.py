"""団体戦PDF/HTMLパーサー: 関東連盟マトリクス形式から立教大学の成績を抽出する。"""

import re
from collections import defaultdict
from pathlib import Path

import pdfplumber
from bs4 import BeautifulSoup

from common import decode_html, is_rikkyo, logger

PROMOTION_KEYWORDS = {"昇級", "降級", "優勝", "準優勝"}
ROUND_ROBIN_HEADERS = re.compile(r"勝点|勝数|順位")


def _group_words_by_line(page, tolerance: int = 5) -> list[list[str]]:
    """ページの単語を y 座標でグルーピングして行リストに変換する。"""
    words = page.extract_words()
    rows: dict[int, list[str]] = defaultdict(list)
    for w in words:
        key = round(w["top"] / tolerance) * tolerance
        rows[key].append(w["text"])
    return [rows[k] for k in sorted(rows.keys())]


def _parse_rikkyo_row(tokens: list[str]) -> dict | None:
    """
    立教を含む行トークンから成績を抽出する。
    形式: [seeding] [name] [N-1 opponent scores] [勝数] [勝点] [順位] [入れ替え?]
    右端から解析する。「8降」のような複合トークンや浮動小数点ゴミにも対応。
    """
    promotion = None

    # 末尾の整数でないゴミ(浮動小数点・座標等)を除去
    # ただし「8降」「1昇」のような複合トークンは有効
    while tokens:
        last = tokens[-1]
        if (re.match(r"^\d+$", last)
                or last in PROMOTION_KEYWORDS
                or re.match(r"^[昇降優準]", last)
                or re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", last)):
            break
        tokens = tokens[:-1]

    if not tokens:
        return None

    # 右端が昇降級ワードなら取る
    if tokens[-1] in PROMOTION_KEYWORDS:
        promotion = tokens[-1]
        tokens = tokens[:-1]
    # 「8降」「1昇」のような複合トークンを分離
    elif re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", tokens[-1]):
        m = re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", tokens[-1])
        promotion = "昇級" if "昇" in m.group(2) else "降級"
        tokens = tokens[:-1] + [m.group(1)]

    # 以降: [...scores...] 勝数 勝点 順位
    if len(tokens) < 3:
        return None

    try:
        rank = int(tokens[-1])
        points = int(tokens[-2])    # 勝点(ボード合計)
        wins_matches = int(tokens[-3])  # 勝数(試合勝数)
    except (ValueError, IndexError):
        return None

    return {
        "rank": rank,
        "wins": wins_matches,
        "losses": None,
        "points": points,
        "promotion": promotion,
        "note": "",
    }


DIVISION_RE = re.compile(
    r"([A-ZＡ-Ｚ][0-9０-９]?(?:\s*[0-9０-９])?[級])"
)


def _normalize_division(div: str) -> str:
    """全角文字の級名を半角に正規化する。"""
    div = div.translate(str.maketrans("ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃ０１２３４５６７８９　",
                                       "ABCDEFGHIJKLMNOPQRSTUVWXYZabc0123456789 "))
    return div.strip()


def _find_division(lines: list[list[str]], rikkyo_line_idx: int) -> str | None:
    """立教行の上方向に走査して直近の級名(A級/B2級等)を探す。半角・全角両対応。"""
    for i in range(rikkyo_line_idx - 1, -1, -1):
        text = " ".join(lines[i])
        m = DIVISION_RE.search(text)
        if m:
            return _normalize_division(m.group(1))
    return None


def parse_team_pdf(pdf_path: Path, event_name: str) -> dict | None:
    """
    団体戦 PDF を解析し、立教の成績を返す。
    立教が出場していない場合は None を返す。
    """
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = _group_words_by_line(page)
            for idx, tokens in enumerate(lines):
                line_text = " ".join(tokens)
                if not is_rikkyo(line_text):
                    continue
                # ヘッダー行(列名として「立教」が出る)はスキップ
                if "勝点" in tokens or "勝数" in tokens:
                    continue

                result = _parse_rikkyo_row(tokens[:])
                if result is None:
                    logger.warning("立教行の解析失敗: %s", line_text)
                    continue

                division = _find_division(lines, idx)
                logger.info("立教発見: %s %s rank=%s wins=%s points=%s promotion=%s",
                            event_name, division or "?",
                            result["rank"], result["wins"], result["points"], result["promotion"])
                return {"division": division, "result": result}

    return None


def parse_team_html(html_path: Path, event_name: str) -> dict | None:
    """
    団体戦 HTML(Excel 由来の表)を解析し、立教の成績を返す。
    ＼ が対角線マーカー。右端から 順位・勝点・勝数 を読む。
    """
    raw = html_path.read_bytes()
    html = decode_html(raw)
    soup = BeautifulSoup(html, "html.parser")

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        last_header_cells: list[str] = []
        division = None
        for row in rows:
            cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if not cells:
                continue

            # ヘッダー行(勝点/勝数を含む行)を記憶
            joined = " ".join(cells)
            if "勝点" in joined or "勝数" in joined:
                last_header_cells = cells
                m = DIVISION_RE.search(joined)
                if m:
                    division = _normalize_division(m.group(1))
                continue

            # 立教行かチェック
            if not is_rikkyo(joined):
                continue

            # 立教大学が見つかった列インデックスを特定
            rikkyo_idx = next(
                (i for i, c in enumerate(cells) if is_rikkyo(c)), 0
            )
            relevant_cells = cells[rikkyo_idx:]  # 立教大学セルから右側

            # ヘッダー行で立教列より手前にある最後の division ラベルを探す
            if last_header_cells:
                best_div = None
                for col_i, hcell in enumerate(last_header_cells):
                    m = DIVISION_RE.match(hcell.strip())
                    if m and col_i <= rikkyo_idx:
                        best_div = _normalize_division(m.group(1))
                if best_div:
                    division = best_div

            # ＼(対角線)を除去して数値のみ抽出
            nums = [c for c in relevant_cells[1:] if c and c not in ("＼", "\\", "／", "")]
            try:
                # 順位セルが「2昇」「3降」のように promotion を含む場合を分離
                promotion = None
                rank_raw = None
                clean_nums = []
                for n in nums:
                    if re.match(r"^\d+$", n):
                        clean_nums.append(int(n))
                    elif re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)", n):
                        m_p = re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)", n)
                        rank_raw = int(m_p.group(1))
                        promo_char = m_p.group(2)
                        promotion = "昇級" if "昇" in promo_char else "降級"
                    elif n in PROMOTION_KEYWORDS:
                        promotion = n

                if rank_raw is not None:
                    # 順位が別セル(rank_raw)に入っている場合:
                    # clean_nums は [...scores, 勝数, 勝点]
                    rank = rank_raw
                    points = clean_nums[-1] if clean_nums else None
                    wins_matches = clean_nums[-2] if len(clean_nums) >= 2 else None
                else:
                    # 順位が clean_nums の末尾にある場合:
                    # clean_nums は [...scores, 勝数, 勝点, 順位]
                    rank = clean_nums[-1]
                    points = clean_nums[-2] if len(clean_nums) >= 2 else None
                    wins_matches = clean_nums[-3] if len(clean_nums) >= 3 else None
            except (ValueError, IndexError):
                logger.warning("HTML立教行の解析失敗: %s", cells[:8])
                continue

            logger.info("立教発見(HTML): %s %s rank=%s wins=%s points=%s",
                        event_name, division or "?", rank, wins_matches, points)
            return {"division": division, "result": {
                "rank": rank,
                "wins": wins_matches,
                "losses": None,
                "points": points,
                "promotion": promotion,
                "note": "",
            }}

    return None


STATS_TOKENS = frozenset({"勝点", "勝数", "順位", "入れ替え", "入替え", "入替"})


def _to_float_val(s: str | None) -> float | None:
    """文字列を float に変換する。空文字・対角線記号は None を返す。"""
    if s is None or s == "" or s in ("＼", "\\", "/", "／"):
        return None
    s2 = str(s).translate(str.maketrans("０１２３４５６７８９．ー－", "0123456789.--"))
    try:
        return float(s2.strip())
    except ValueError:
        return None


def _to_int_val(s: str | None) -> int | None:
    if s is None or s == "":
        return None
    s2 = str(s).translate(str.maketrans("０１２３４５６７８９", "0123456789"))
    try:
        return int(s2.strip())
    except ValueError:
        return None


def _parse_promotion(raw: str) -> tuple[int | None, str | None]:
    """'8降' '1昇' '3' のようなランク文字列を (rank, promotion) に分解する。"""
    m = re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)?$", str(raw).strip())
    if not m:
        return None, None
    rank = int(m.group(1))
    promo_char = m.group(2) or ""
    if "昇" in promo_char:
        promotion = "昇級"
    elif "降" in promo_char:
        promotion = "降級"
    else:
        promotion = None
    return rank, promotion


def _build_table_row_pdf(tokens: list, seeding: int, n_teams: int) -> dict | None:
    """PDF行トークン(seeding・name 除去済み)から round-robin 行を構築する。"""
    values = list(tokens)
    promotion = None
    rank_embedded = None

    # 末尾の座標アーティファクト(入れ替え列等の大きな数値)を除去
    # 例: '3020.04', '7036.07', '7028' → 有効値の最大は勝点≈50程度なので1000以上は不要
    while values:
        try:
            if float(str(values[-1])) >= 1000:
                values.pop()
                continue
        except (ValueError, TypeError):
            pass
        break

    # 末尾の非純数値トークンを処理
    # パターン1: '降級'/'昇級' (テキストのみ) → promotionのみ、rankは次トークン
    # パターン2: '2昇'/'8降' (数字+昇降) → rankとpromotionが結合
    if values and not re.match(r"^[\d.]+$", str(values[-1])):
        last = str(values.pop())
        m = re.match(r"^(\d+)(昇|降|昇級|降級|優勝|準優勝)$", last)
        if m:
            rank_embedded = int(m.group(1))
            promotion = "昇級" if "昇" in m.group(2) else "降級"
        elif "昇" in last:
            promotion = "昇級"
        elif "降" in last:
            promotion = "降級"

    if rank_embedded is not None:
        # ランクが末尾トークンに埋め込まれていた場合: 残りは scores + wins + points
        if len(values) < 2:
            return None
        points = _to_float_val(str(values[-1]))
        wins = _to_float_val(str(values[-2])) if len(values) >= 3 else None
        raw_scores = [_to_float_val(str(v)) for v in values[:-2]] if len(values) >= 3 else [_to_float_val(str(v)) for v in values[:-1]]
        rank = rank_embedded
    else:
        # 通常形式: 右端3つが wins, points, rank
        if len(values) < 3:
            return None
        rank_raw, points_raw, wins_raw = str(values[-1]), str(values[-2]), str(values[-3])
        rank, promo2 = _parse_promotion(rank_raw)
        if promo2:
            promotion = promo2
        points = _to_float_val(points_raw)
        wins = _to_float_val(wins_raw)
        raw_scores = [_to_float_val(str(v)) for v in values[:-3]]

    # PDF: 対角線セルなし → seeding-1 の位置に None を挿入
    diag = seeding - 1
    full_scores = raw_scores[:diag] + [None] + raw_scores[diag:]
    return {"seeding": seeding, "scores": full_scores, "wins": wins, "points": points,
            "rank": rank, "promotion": promotion}


def _build_table_row_html(cells: list, seeding: int, n_teams: int) -> dict | None:
    """HTML セルリスト(seeding・name 除去済み)から round-robin 行を構築する。"""
    raw_scores_str = cells[:n_teams]
    # 末尾の空文字を除去してから stats を処理する
    stats = [s for s in cells[n_teams:] if str(s).strip() != ""]
    if len(stats) < 2:
        return None
    rank, promotion = _parse_promotion(str(stats[-1]))
    points = _to_float_val(str(stats[-2])) if len(stats) >= 2 else None
    wins = _to_float_val(str(stats[-3])) if len(stats) >= 3 else None
    full_scores: list = []
    for s in raw_scores_str:
        sv = str(s)
        if sv == "" or sv in ("＼", "\\", "/", "／"):
            full_scores.append(None)
        else:
            full_scores.append(_to_float_val(sv))
    return {"seeding": seeding, "scores": full_scores, "wins": wins, "points": points,
            "rank": rank, "promotion": promotion}


def parse_full_table_pdf(pdf_path: Path, target_division: str | None = None) -> dict | None:
    """
    PDF から全チーム総当たりテーブルを抽出する。
    target_division: 'B2級' 等。None の場合は立教所属の級を自動選択。
    """
    tables: dict[str, dict] = {}
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            lines = _group_words_by_line(page)
            i = 0
            while i < len(lines):
                tokens = lines[i]
                if "勝点" not in tokens or "順位" not in tokens or len(tokens) < 4:
                    i += 1
                    continue
                stats_start = next((j for j, t in enumerate(tokens) if t in STATS_TOKENS), len(tokens))
                div_token = tokens[0]
                if not re.search(r"[A-ZＡ-Ｚ]\d?[級]", div_token):
                    i += 1
                    continue
                division = _normalize_division(div_token)
                team_abbrevs = tokens[1:stats_start]
                n_teams = len(team_abbrevs)

                section_rows: list[dict] = []
                j = i + 1
                implicit_seed = 1  # for PDFs without an explicit seeding column
                while j < len(lines):
                    row = lines[j]
                    if not row:
                        break
                    if any(t in STATS_TOKENS for t in row):
                        break
                    # Check for next division header
                    if re.search(r"[A-ZＡ-Ｚ]\d?[級]", str(row[0])) and len(row) >= 4:
                        break
                    first = str(row[0])
                    if re.match(r"^\d+$", first):
                        # Explicit seeding column
                        seeding = int(first)
                        team = str(row[1])
                        rest = row[2:]
                    else:
                        # No seeding column (e.g. H27 format) — use implicit position
                        seeding = implicit_seed
                        team = first
                        rest = row[1:]
                    implicit_seed += 1
                    built = _build_table_row_pdf(list(rest), seeding, n_teams)
                    if built:
                        section_rows.append({"team": team, **built})
                    j += 1

                if len(section_rows) >= 3:
                    section_rows.sort(key=lambda r: r["seeding"])
                    tables[division] = {
                        "division": division,
                        "teams": [r["team"] for r in section_rows],
                        "team_abbrevs": team_abbrevs,
                        "rows": section_rows,
                    }
                i = j

    if not tables:
        return None
    if target_division and target_division in tables:
        return tables[target_division]
    # 立教が所属する級を選択
    for tbl in tables.values():
        if any(is_rikkyo(r["team"]) for r in tbl["rows"]):
            return tbl
    return next(iter(tables.values()))


def parse_full_table_html(html_path: Path, target_division: str | None = None) -> dict | None:
    """
    HTML(関東連盟 Excel 由来)から全チーム総当たりテーブルを抽出する。
    複数級が横並びの形式に対応。
    """
    raw = html_path.read_bytes()
    html = decode_html(raw)
    soup = BeautifulSoup(html, "html.parser")

    tables_found: dict[str, dict] = {}

    for table in soup.find_all("table"):
        rows_list = table.find_all("tr")
        for row_idx, header_row in enumerate(rows_list):
            header_cells = [td.get_text(strip=True) for td in header_row.find_all(["td", "th"])]
            if not header_cells:
                continue

            # 級ラベルの位置を検出
            div_positions: list[tuple[int, str]] = []
            for col_idx, cell in enumerate(header_cells):
                norm = _normalize_division(cell)
                if re.search(r"[A-Z]\d?級", norm):
                    div_positions.append((col_idx, norm))
            if not div_positions:
                continue

            div_positions.sort(key=lambda x: x[0])

            for sec_idx, (sec_start, division_label) in enumerate(div_positions):
                sec_end = div_positions[sec_idx + 1][0] if sec_idx + 1 < len(div_positions) else len(header_cells)

                # stats 列を探す
                stats_col = next(
                    (c for c in range(sec_start, sec_end) if header_cells[c] in STATS_TOKENS),
                    None,
                )
                if stats_col is None:
                    continue

                # 級ラベルの直後にあるシード列ヘッダー('')や大学名列ヘッダー('大学名')をスキップ
                abbrev_start = sec_start + 1
                _TEAM_COL_HEADERS = {"", "大学名", "チーム", "対校"}
                while (abbrev_start < sec_start + 3
                       and abbrev_start < len(header_cells)
                       and header_cells[abbrev_start] in _TEAM_COL_HEADERS):
                    abbrev_start += 1
                team_abbrevs = header_cells[abbrev_start:stats_col]
                n_teams = len(team_abbrevs)
                if n_teams < 3:
                    continue

                section_rows: list[dict] = []
                seen_seeds: set[int] = set()
                for data_row in rows_list[row_idx + 1:]:
                    data_cells = [td.get_text(strip=True) for td in data_row.find_all(["td", "th"])]
                    if len(data_cells) <= sec_start:
                        continue
                    cell_at_start = data_cells[sec_start] if sec_start < len(data_cells) else ""
                    # Stop at next division header
                    if re.search(r"[A-ZＡ-Ｚ]\d?[級]", _normalize_division(cell_at_start)):
                        break
                    section = data_cells[sec_start:sec_end]
                    if re.match(r"^\d+$", str(cell_at_start)):
                        # 明示的シード: [seeding, team, scores..., stats...]
                        seeding = int(cell_at_start)
                        if seeding in seen_seeds:
                            break
                        seen_seeds.add(seeding)
                        if len(section) < 2:
                            continue
                        team = str(section[1])
                        rest = section[2:]
                    elif cell_at_start.strip():
                        # 暗黙シード: [team, scores..., stats...] (H26秋等)
                        seeding = len(section_rows) + 1
                        if seeding in seen_seeds:
                            break
                        seen_seeds.add(seeding)
                        team = str(cell_at_start)
                        rest = section[1:]
                    else:
                        continue
                    built = _build_table_row_html(rest, seeding, n_teams)
                    if built:
                        section_rows.append({"team": team, **built})

                if len(section_rows) >= 3:
                    section_rows.sort(key=lambda r: r["seeding"])
                    tables_found[division_label] = {
                        "division": division_label,
                        "teams": [r["team"] for r in section_rows],
                        "team_abbrevs": team_abbrevs,
                        "rows": section_rows,
                    }

    if not tables_found:
        return None
    if target_division and target_division in tables_found:
        return tables_found[target_division]
    for tbl in tables_found.values():
        if any(is_rikkyo(r["team"]) for r in tbl["rows"]):
            return tbl
    return next(iter(tables_found.values()))


def parse_full_table_xlsx(xlsx_path: Path, target_division: str | None = None) -> dict | None:
    """
    XLSX(関東連盟 Excel)から全チーム総当たりテーブルを抽出する。
    対応形式: チーム名が先頭列・明示的シードなし(H27春等)
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl がインストールされていません")
        return None

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    tables_found: dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        i = 0
        while i < len(all_rows):
            raw = all_rows[i]
            cells = [str(c) if c is not None else "" for c in raw]

            # ヘッダー行検出: 先頭セルが級ラベル、かつ '勝点' か '勝数' を含む
            first = cells[0].strip() if cells else ""
            norm_first = _normalize_division(first)
            has_stats = any(c in STATS_TOKENS for c in cells)
            if re.search(r"[A-Z]\d?[級]", norm_first) and has_stats:
                division = norm_first
                # stats 列の開始位置
                stats_col = next((j for j, c in enumerate(cells) if c in STATS_TOKENS), len(cells))
                # 略称は division の次から stats 列の直前まで
                team_abbrevs = [c for c in cells[1:stats_col] if c.strip()]
                n_teams = len(team_abbrevs)
                if n_teams < 3:
                    i += 1
                    continue

                section_rows: list[dict] = []
                j = i + 1
                while j < len(all_rows):
                    row_raw = all_rows[j]
                    row_cells = [str(c) if c is not None else "" for c in row_raw]
                    row_first = row_cells[0].strip() if row_cells else ""
                    if not row_first:
                        j += 1
                        continue
                    # 次の級ヘッダーで終了
                    if re.search(r"[A-Z]\d?[級]", _normalize_division(row_first)) and any(
                        c in STATS_TOKENS for c in row_cells
                    ):
                        break
                    # データ行: チーム名が先頭(暗黙シード)
                    seeding = len(section_rows) + 1
                    team = row_first
                    # scores + stats (trailing None/'' を除去)
                    rest = [c for c in row_cells[1:] if c != "None"]
                    # 末尾の空セルを除去
                    while rest and rest[-1].strip() == "":
                        rest.pop()
                    built = _build_table_row_html(rest, seeding, n_teams)
                    if built:
                        section_rows.append({"team": team, **built})
                    j += 1

                if len(section_rows) >= 3:
                    section_rows.sort(key=lambda r: r["seeding"])
                    tables_found[division] = {
                        "division": division,
                        "teams": [r["team"] for r in section_rows],
                        "team_abbrevs": team_abbrevs,
                        "rows": section_rows,
                    }
                i = j
            else:
                i += 1

    if not tables_found:
        return None
    if target_division and target_division in tables_found:
        return tables_found[target_division]
    for tbl in tables_found.values():
        if any(is_rikkyo(r["team"]) for r in tbl["rows"]):
            return tbl
    return next(iter(tables_found.values()))


def _parse_schedule_text(text: str) -> tuple[int, str | None, str | None]:
    """'1日目 4月27日 於 東京理科大学' のようなテキストから (day, date, venue) を抽出する。"""
    text = text.strip().replace("　", " ")
    m = re.match(r"(\d+)日目", text)
    if not m:
        return 0, None, None
    day = int(m.group(1))
    date_m = re.search(r"((?:平成|令和)\d+年\d+月\d+日[（(][日月火水木金土][)）]?|\d{4}/\d+/\d+|\d+月\d+日)", text)
    date = date_m.group(1) if date_m else None
    venue_m = re.search(r"於[・\s]*(.+)$", text)
    venue = venue_m.group(1).strip() if venue_m else None
    return day, date, venue


def parse_schedule_from_html(html_path: Path) -> list[dict]:
    """HTML の先頭付近から複数日程情報を抽出する。"""
    raw = html_path.read_bytes()
    html = decode_html(raw)
    soup = BeautifulSoup(html, "html.parser")
    schedule: dict[int, dict] = {}

    for table in soup.find_all("table"):
        rows = table.find_all("tr")
        for row in rows[:12]:
            cells = [c.get_text(strip=True) for c in row.find_all(["td", "th"])]
            row_text = " ".join(cells)
            if "日目" not in row_text:
                continue
            # パターン1: 一つのセルに '1日目 4月27日 於 東京理科大学' が入っている
            for cell in cells:
                if re.match(r"\d+日目", cell):
                    day, date, venue = _parse_schedule_text(cell)
                    if day > 0 and day not in schedule:
                        schedule[day] = {"day": day, "date": date, "venue": venue}
            # パターン2: ['1日目', '2011/10/16', '', '於', '東京農業大学'] のように分かれている
            for idx, cell in enumerate(cells):
                if re.match(r"^\d+日目$", cell):
                    day = int(re.match(r"(\d+)", cell).group(1))
                    if day in schedule:
                        continue
                    date = cells[idx + 1] if idx + 1 < len(cells) else None
                    venue_parts = [c for c in cells[idx + 2:] if c and c != "於"]
                    venue = " ".join(venue_parts).lstrip("・ ").strip() or None
                    schedule[day] = {"day": day, "date": date, "venue": venue}
        if schedule:
            break
    return [schedule[k] for k in sorted(schedule)]


def parse_schedule_from_xlsx(xlsx_path: Path) -> list[dict]:
    """XLSX の先頭付近から複数日程情報を抽出する。"""
    try:
        import openpyxl
    except ImportError:
        return []
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    schedule: dict[int, dict] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 10:
                break
            for val in row:
                if val is None:
                    continue
                text = str(val).strip()
                if re.match(r"\d+日目", text):
                    day, date, venue = _parse_schedule_text(text)
                    if day > 0 and day not in schedule:
                        schedule[day] = {"day": day, "date": date, "venue": venue}
    return [schedule[k] for k in sorted(schedule)]


def extract_team_losses(result: dict, total_teams: int | None = None) -> dict:
    """勝数と総チーム数から負数を補完する(可能な場合)。"""
    r = result.copy()
    if r.get("wins") is not None and total_teams is not None:
        r["losses"] = (total_teams - 1) - r["wins"]
    return r
